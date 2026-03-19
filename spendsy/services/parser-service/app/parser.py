# pyright: reportMissingImports=false
# pyright: reportCallIssue=false
# pyright: reportArgumentType=false
# pyright: reportIndexIssue=false
# pyright: reportOperatorIssue=false
# pyright: reportAttributeAccessIssue=false
# pyre-ignore-all-errors

from __future__ import annotations

import csv
import io
import json
import logging
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Literal

import httpx
from tenacity import retry, stop_after_attempt, wait_exponential

import pdfplumber  # type: ignore
from pydantic import BaseModel, ConfigDict, Field  # type: ignore

logger = logging.getLogger(__name__)

try:
    import polars as pl  # type: ignore
except Exception:  # pragma: no cover
    pl = None  # type: ignore[assignment]

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover
    openpyxl = None  # type: ignore[assignment]

try:
    import pytesseract  # type: ignore
except Exception:  # pragma: no cover
    pytesseract = None  # type: ignore[assignment]

DATE_CANDIDATE = re.compile(r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b")
AMOUNT_CANDIDATE = re.compile(r"-?(?:\d{1,3}(?:,\d{2,3})+|\d+)(?:\.\d{1,2})?")
OCR_NUMERIC = re.compile(r"\d+\.\d{2}")
MONEY_LIKE_CELL = re.compile(
    r"^\s*(?:[+-]|\()?\s*(?:₹|INR|RS\.?)?\s*(?:\d{1,3}(?:,\d{2,3})+|\d+)(?:\.\d{1,2})?\s*\)?\s*$",
    re.IGNORECASE,
)
SUMMARY_PATTERNS = (
    re.compile(r"^\s*(opening|closing)\s+balance\b", re.IGNORECASE),
    re.compile(r"^\s*(b/f|brought forward)\b", re.IGNORECASE),
    re.compile(r"^\s*total\s+(debit|credit|withdraw|deposit|amount|txn|transactions)\b", re.IGNORECASE),
)
DATE_ALIASES = ("date", "txn date", "transaction date", "value date", "tran date")
DESCRIPTION_ALIASES = (
    "description",
    "narration",
    "particulars",
    "details",
    "remarks",
    "transaction remarks",
    "transaction details",
    "ref particulars",
)
DEBIT_ALIASES = ("debit", "withdrawal", "dr", "amt out", "payment", "withdrawal amt", "debit amount")
CREDIT_ALIASES = ("credit", "deposit", "cr", "amt in", "receipt", "deposit amt", "credit amount")
AMOUNT_ALIASES = ("amount", "txn amount", "transaction amount")
BALANCE_ALIASES = ("balance", "running balance", "closing balance", "available balance")


def detect_bank(first_page_text: str) -> str:
    text = (first_page_text or "").lower()

    if "icici bank" in text or re.search(r"\bicici\b", text):
        return "icici"
    if "hdfc bank" in text or re.search(r"\bhdfc\b", text):
        return "hdfc"
    if "state bank of india" in text or re.search(r"\bsbi\b", text):
        return "sbi"
    if "axis bank" in text or re.search(r"\baxis\b", text):
        return "axis"
    return "generic"


class ParsedTransaction(BaseModel):
    model_config = ConfigDict(extra="forbid")

    date: date
    description: str
    amount: float
    type: Literal["income", "expense"]
    debit: float | None = None
    credit: float | None = None
    balance: float | None = None
    confidence: float | None = None
    source: Literal["statement"] = "statement"
    is_valid: bool = True


class ParserResponse(BaseModel):
    model_config = ConfigDict(extra="forbid")

    status: str
    reconciliation_score: float
    transactions: list[ParsedTransaction]
    meta: dict[str, Any] = Field(default_factory=dict)


@dataclass(frozen=True)
class _HeaderMap:
    date_idx: int
    desc_idx: int
    debit_idx: int | None
    credit_idx: int | None
    amount_idx: int | None
    balance_idx: int | None


@dataclass(frozen=True)
class _OCRLine:
    text: str
    confidence: float
    top: int
    left: int
    tokens: tuple[tuple[int, str], ...] = ()


@dataclass(frozen=True)
class _OCRColumnHints:
    expense_x: float | None = None
    income_x: float | None = None
    balance_x: float | None = None


class IntegratedParser:
    def __init__(self) -> None:
        pass

    def parse(self, file_bytes: bytes, filename: str | None = None, content_type: str | None = None) -> ParserResponse:
        parsed_rows: list[ParsedTransaction] = []
        method = "digital"
        file_type = "pdf"
        bank = self._detect_bank_from_input(file_bytes, filename, content_type)

        if self._is_csv_input(file_bytes, filename, content_type):
            file_type = "csv"
            parsed_rows = self._extract_csv(file_bytes, bank=bank)
            method = "csv"
        elif self._is_xlsx_input(file_bytes, filename, content_type):
            file_type = "xlsx"
            parsed_rows = self._extract_xlsx(file_bytes, bank=bank)
            method = "xlsx"
        else:
            digital_rows = self._extract_digital(file_bytes, bank=bank)
            parsed_rows = digital_rows
            method = "digital"

            # Hybrid fallback: if digital extraction is incomplete (< 90% confidence), supplement with OCR rows.
            # This handles cases where pdfplumber misses some rows or where tables are poorly formatted.
            ocr_rows = self._extract_ocr(file_bytes, bank=bank)
            if ocr_rows and len(digital_rows) < len(ocr_rows):
                # OCR found more transactions than digital; use OCR or hybrid approach
                merged: dict[tuple[str, str, int, str], ParsedTransaction] = {
                    self._tx_key(tx): tx for tx in digital_rows
                }
                for tx in ocr_rows:
                    merged.setdefault(self._tx_key(tx), tx)
                parsed_rows = list(merged.values())
                method = "ocr" if not digital_rows else "hybrid"

        if file_type in {"csv", "xlsx"} and not parsed_rows:
            return ParserResponse(
                status="no_transactions",
                reconciliation_score=1.0,
                transactions=[],
                meta={"bank": bank, "method": method, "checksum_verified": True, "count": 0},
            )

        cleaned = self._clean_with_polars(parsed_rows)
        reconciled, score = self.verify_integrity(cleaned)
        ordered = sorted(
            reconciled,
            key=lambda t: (t.date.isoformat(), t.description.lower(), round(float(t.amount), 2), t.type),
        )
        confidences = [float(tx.confidence) for tx in ordered if tx.confidence is not None]
        avg_confidence = round(sum(confidences) / len(confidences), 4) if confidences else (0.99 if method == "digital" else 0.9)
        min_confidence = round(min(confidences), 4) if confidences else avg_confidence
        requires_review = method in {"ocr", "hybrid"} and (
            avg_confidence < 0.85
            or min_confidence < 0.7
            or score < 0.9
            or any(not tx.is_valid for tx in ordered)
        )
        return ParserResponse(
            status="success" if ordered else "no_transactions",
            reconciliation_score=round(score, 4),
            transactions=ordered,
            meta={
                "bank": bank,
                "method": method,
                "avg_confidence": avg_confidence,
                "min_confidence": min_confidence,
                "requires_review": requires_review,
                "checksum_verified": bool(score >= 0.9 and avg_confidence >= 0.85 if ordered else True),
                "count": len(ordered),
            },
        )


    def _detect_bank_from_input(self, file_bytes: bytes, filename: str | None, content_type: str | None) -> str:
        name = filename or ""
        if self._is_csv_input(file_bytes, filename, content_type):
            return detect_bank(f"{name}\n{self._decode_text(file_bytes[:4096])}")
        if self._is_xlsx_input(file_bytes, filename, content_type):
            return detect_bank(f"{name}\n{self._preview_xlsx_text(file_bytes)}")
        return self._detect_bank_from_pdf(file_bytes, filename)

    def _detect_bank_from_pdf(self, file_bytes: bytes, filename: str | None = None) -> str:
        preview = filename or ""
        try:
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                if pdf.pages:
                    preview = f"{preview}\n{pdf.pages[0].extract_text() or ''}"
        except Exception:
            return detect_bank(preview)
        return detect_bank(preview)

    def _preview_xlsx_text(self, file_bytes: bytes) -> str:
        if openpyxl is None:
            return ""
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception:
            return ""

        snippets: list[str] = []
        for sheet in workbook.worksheets[:1]:
            for row in sheet.iter_rows(values_only=True, max_row=8):
                parts = [self._normalize_cell(cell) for cell in row if self._normalize_cell(cell)]
                if parts:
                    snippets.append(" ".join(parts))
        return "\n".join(snippets)

    def _extract_digital(self, pdf_bytes: bytes, bank: str = "generic") -> list[ParsedTransaction]:
        transactions: list[ParsedTransaction] = []
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page_idx, page in enumerate(pdf.pages):
                tables = page.extract_tables() or []
                logger.info("parser_stage=digital_extract page=%d tables_found=%d", page_idx + 1, len(tables))
                for table in tables:
                    transactions.extend(self._parse_structured_table(table, bank=bank))
        logger.info("parser_stage=digital_parse bank=%s transactions=%d", bank, len(transactions))
        return transactions

    def _extract_xlsx(self, file_bytes: bytes, bank: str = "generic") -> list[ParsedTransaction]:
        if openpyxl is None:
            logger.warning("parser_stage=xlsx_extract skipped=missing_openpyxl")
            return []

        transactions: list[ParsedTransaction] = []
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception:
            logger.warning("parser_stage=xlsx_extract invalid_workbook")
            return []
        for sheet in workbook.worksheets:
            rows = [
                [self._normalize_cell(cell) for cell in row]
                for row in sheet.iter_rows(values_only=True)
                if any(self._normalize_cell(cell) for cell in row)
            ]
            if not rows:
                continue
            transactions.extend(self._parse_structured_table(rows, bank=bank))
        logger.info("parser_stage=xlsx_parse bank=%s transactions=%d", bank, len(transactions))
        return transactions

    def _extract_csv(self, file_bytes: bytes, bank: str = "generic") -> list[ParsedTransaction]:
        rows: list[dict[str, str]] = []

        if pl is not None:
            try:
                df = pl.read_csv(io.BytesIO(file_bytes), ignore_errors=True, infer_schema_length=5000)
                rows = [{k: str(v or "") for k, v in row.items()} for row in df.iter_rows(named=True)]
            except Exception:
                rows = []

        if not rows:
            text = self._decode_text(file_bytes)
            if not text:
                return []
            reader = csv.DictReader(io.StringIO(text))
            rows = [{str(k or ""): str(v or "") for k, v in row.items()} for row in reader]

        parsed: list[ParsedTransaction] = []
        for row in rows:
            tx = self._parse_csv_row(row, bank=bank)
            if tx is not None:
                parsed.append(tx)

        logger.info("parser_stage=csv_parse bank=%s rows=%d transactions=%d", bank, len(rows), len(parsed))
        return parsed

    def _extract_ocr(self, pdf_bytes: bytes, bank: str = "generic") -> list[ParsedTransaction]:
        if pytesseract is None:
            logger.warning("parser_stage=ocr_extract skipped=missing_pytesseract")
            return []

        lines: list[_OCRLine] = []
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page_idx, page in enumerate(pdf.pages[:10]):
                image = page.to_image(resolution=220).original
                page_lines = self._extract_ocr_lines_from_image(image)
                logger.info("parser_stage=ocr_extract page=%d lines=%d", page_idx + 1, len(page_lines))
                lines.extend(page_lines)

        items, skipped, valid = self._parse_ocr_lines(lines, bank=bank)
        logger.info(
            "parser_stage=ocr_parse bank=%s total_lines=%d valid_rows=%d skipped_rows=%d transactions=%d",
            bank,
            len(lines),
            valid,
            skipped,
            len(items),
        )
        return items

    def _extract_ocr_lines_from_image(self, image) -> list[_OCRLine]:
        if pytesseract is None:
            return []

        try:
            data = pytesseract.image_to_data(image, config="--oem 3 --psm 6", output_type=pytesseract.Output.DICT)
            lines = self._ocr_lines_from_data(data)
            if lines:
                return lines
        except Exception:
            logger.warning("parser_stage=ocr_extract image_to_data_failed", exc_info=True)

        text = pytesseract.image_to_string(image, config="--oem 3 --psm 6")
        return [
            _OCRLine(text=self._normalize_cell(x), confidence=0.45, top=idx * 20, left=0)
            for idx, x in enumerate(text.splitlines())
            if self._normalize_cell(x)
        ]

    def _ocr_lines_from_data(self, data: dict[str, list[Any]]) -> list[_OCRLine]:
        texts = data.get("text", [])
        if not texts:
            return []

        grouped: dict[tuple[int, int, int, int], dict[str, Any]] = {}
        count = len(texts)
        for idx in range(count):
            raw_text = self._normalize_cell(texts[idx])
            if not raw_text:
                continue
            conf = self._safe_ocr_conf(data.get("conf", [])[idx] if idx < len(data.get("conf", [])) else None)
            if conf <= 0:
                continue

            key = (
                int(self._safe_ocr_number(data.get("page_num", [])[idx] if idx < len(data.get("page_num", [])) else 0)),
                int(self._safe_ocr_number(data.get("block_num", [])[idx] if idx < len(data.get("block_num", [])) else 0)),
                int(self._safe_ocr_number(data.get("par_num", [])[idx] if idx < len(data.get("par_num", [])) else 0)),
                int(self._safe_ocr_number(data.get("line_num", [])[idx] if idx < len(data.get("line_num", [])) else idx)),
            )
            entry = grouped.setdefault(
                key,
                {"tokens": [], "confidences": [], "top": 0, "left": 0},
            )
            left = int(self._safe_ocr_number(data.get("left", [])[idx] if idx < len(data.get("left", [])) else 0))
            top = int(self._safe_ocr_number(data.get("top", [])[idx] if idx < len(data.get("top", [])) else 0))
            entry["tokens"].append((left, raw_text))
            entry["confidences"].append(conf)
            if not entry["top"]:
                entry["top"] = top
            if not entry["left"]:
                entry["left"] = left

        lines: list[_OCRLine] = []
        for key in sorted(grouped.keys()):
            entry = grouped[key]
            tokens = " ".join(text for _, text in sorted(entry["tokens"], key=lambda item: item[0])).strip()
            if not tokens:
                continue
            confidences = entry["confidences"] or [0.0]
            lines.append(
                _OCRLine(
                    text=tokens,
                    confidence=round(sum(confidences) / len(confidences) / 100.0, 4),
                    top=int(entry["top"]),
                    left=int(entry["left"]),
                )
            )
        return lines

    def _safe_ocr_conf(self, value: Any) -> float:
        try:
            return max(float(value), 0.0)
        except (TypeError, ValueError):
            return 0.0

    def _safe_ocr_number(self, value: Any) -> float:
        try:
            return float(value)
        except (TypeError, ValueError):
            return 0.0

    def _parse_structured_table(self, table: list[list[str | None]], bank: str = "generic") -> list[ParsedTransaction]:
        if not table:
            return []
        header = self._detect_header(table)
        if header is None:
            return []

        transactions: list[ParsedTransaction] = []
        for row in self._coalesce_table_rows(table[1:], header, bank=bank):
            tx = self._parse_table_row(row, header, bank=bank)
            if tx is not None:
                transactions.append(tx)
        return transactions

    def _detect_header(self, table: list[list[str | None]]) -> _HeaderMap | None:
        for row in table[:4]:
            normalized = [self._normalize_cell(c) for c in row]
            date_idx = self._find_col(normalized, DATE_ALIASES)
            desc_idx = self._find_col(normalized, DESCRIPTION_ALIASES)
            debit_idx = self._find_col(normalized, DEBIT_ALIASES)
            credit_idx = self._find_col(normalized, CREDIT_ALIASES)
            amount_idx = self._find_col(normalized, AMOUNT_ALIASES)
            balance_idx = self._find_col(normalized, BALANCE_ALIASES)
            if date_idx is not None and desc_idx is not None and (amount_idx is not None or debit_idx is not None or credit_idx is not None):
                return _HeaderMap(date_idx, desc_idx, debit_idx, credit_idx, amount_idx, balance_idx)
        return None

    def _find_col(self, headers: list[str], aliases: tuple[str, ...]) -> int | None:
        aliases_norm = [a.lower() for a in aliases]
        for idx, header in enumerate(headers):
            h = header.lower()
            if any(a in h for a in aliases_norm):
                return idx
        return None

    def _coalesce_table_rows(self, rows: list[list[str | None]], hm: _HeaderMap, bank: str = "generic") -> list[list[str | None]]:
        merged: list[list[str | None]] = []
        for raw_row in rows:
            row = list(raw_row or [])
            if merged and self._looks_like_continuation_row(row, merged[-1], hm, bank=bank):
                prev = merged[-1]
                prev_desc = self._cell(prev, hm.desc_idx)
                extra_desc = self._continuation_text(row, hm, bank=bank)
                if extra_desc:
                    if hm.desc_idx >= len(prev):
                        prev.extend([""] * (hm.desc_idx - len(prev) + 1))
                    prev[hm.desc_idx] = f"{prev_desc} {extra_desc}".strip()
                self._carry_forward_row_values(prev, row, hm)
                continue
            merged.append(row)
        return merged

    def _looks_like_continuation_row(self, row: list[str | None], prev: list[str | None], hm: _HeaderMap, bank: str = "generic") -> bool:
        if not self._first_cell_is_empty(row):
            return False
        if self._parse_date(self._cell(row, hm.date_idx)) is not None:
            return False

        desc = self._continuation_text(row, hm, bank=bank)
        if desc and self._is_summary_line(desc):
            return False

        if self._has_primary_transaction_amount(row, hm):
            return False

        if desc:
            return True

        return self._is_balance_only_carry_row(row, prev, hm)

    def _first_cell_is_empty(self, row: list[str | None]) -> bool:
        if not row:
            return True
        return not self._normalize_cell(row[0])

    def _continuation_text(self, row: list[str | None], hm: _HeaderMap, bank: str = "generic") -> str:
        text_indices = {hm.desc_idx}
        if bank == "hdfc":
            text_indices.update(self._description_support_indices(row, hm))
        reserved_indices = {hm.date_idx}
        for idx, value in enumerate(row):
            cell = self._normalize_cell(value)
            if not cell or idx in reserved_indices:
                continue
            text_indices.add(idx)

        parts: list[str] = []
        seen: set[str] = set()
        for idx in sorted(text_indices):
            if idx >= len(row):
                continue
            cell = self._normalize_cell(row[idx])
            if not cell or cell in seen or self._is_monetary_cell(cell):
                continue
            seen.add(cell)
            parts.append(cell)
        return " ".join(parts).strip()

    def _description_support_indices(self, row: list[str | None], hm: _HeaderMap) -> set[int]:
        reserved = {idx for idx in (hm.date_idx, hm.debit_idx, hm.credit_idx, hm.amount_idx, hm.balance_idx) if idx is not None}
        indices: set[int] = set()
        for idx, value in enumerate(row):
            if idx in reserved:
                continue
            cell = self._normalize_cell(value)
            if cell and not self._is_monetary_cell(cell):
                indices.add(idx)
        return indices

    def _has_primary_transaction_amount(self, row: list[str | None], hm: _HeaderMap) -> bool:
        return any(
            idx is not None and self._is_monetary_cell(self._cell(row, idx))
            for idx in (hm.debit_idx, hm.credit_idx, hm.amount_idx)
        )

    def _is_balance_only_carry_row(self, row: list[str | None], prev: list[str | None], hm: _HeaderMap) -> bool:
        if hm.balance_idx is None:
            return False
        balance_value = self._cell(row, hm.balance_idx)
        if not self._is_monetary_cell(balance_value):
            return False
        if self._is_monetary_cell(self._cell(prev, hm.balance_idx)):
            return False
        return bool(self._cell(prev, hm.desc_idx)) and self._has_primary_transaction_amount(prev, hm)

    def _carry_forward_row_values(self, prev: list[str | None], row: list[str | None], hm: _HeaderMap) -> None:
        for idx in (hm.debit_idx, hm.credit_idx, hm.amount_idx, hm.balance_idx):
            if idx is None or idx >= len(row):
                continue
            current = self._cell(row, idx)
            if not current or not self._is_monetary_cell(current):
                continue
            if self._cell(prev, idx):
                continue
            if idx >= len(prev):
                prev.extend([""] * (idx - len(prev) + 1))
            prev[idx] = current  # type: ignore

    def _parse_table_row(self, row: list[str | None], hm: _HeaderMap, bank: str = "generic") -> ParsedTransaction | None:
        raw_date = self._cell(row, hm.date_idx)
        raw_desc = self._row_description(row, hm, bank=bank)
        if not raw_date or not raw_desc:
            return None
        if self._is_summary_line(raw_desc):
            return None

        tx_date = self._parse_date(raw_date)
        if tx_date is None:
            return None

        raw_amount = self._cell(row, hm.amount_idx)
        debit = self._parse_amount(self._cell(row, hm.debit_idx))
        credit = self._parse_amount(self._cell(row, hm.credit_idx))
        amount = self._parse_amount(raw_amount)
        balance = self._parse_amount(self._cell(row, hm.balance_idx))

        tx_type, tx_amount = self._infer_type_and_amount(raw_desc, debit, credit, amount, raw_amount, bank=bank)

        if tx_type is None or tx_amount is None or tx_amount <= 0:
            return None

        return ParsedTransaction(
            date=tx_date,
            description=raw_desc.strip()[:255] or "Transaction",
            amount=tx_amount,
            type=tx_type,
            debit=tx_amount if tx_type == "expense" else None,
            credit=tx_amount if tx_type == "income" else None,
            balance=float(balance) if balance is not None else None,
            is_valid=True,
        )

    def _row_description(self, row: list[str | None], hm: _HeaderMap, bank: str = "generic") -> str:
        if bank != "hdfc":
            return self._cell(row, hm.desc_idx)
        return self._continuation_text(row, hm, bank=bank)

    def _parse_ocr_lines(self, lines: list[_OCRLine] | list[str], bank: str = "generic") -> tuple[list[ParsedTransaction], int, int]:
        items: list[ParsedTransaction] = []
        skipped = 0
        valid = 0

        normalized_lines = [
            line if isinstance(line, _OCRLine) else _OCRLine(text=self._normalize_cell(line), confidence=0.45, top=idx * 20, left=0)
            for idx, line in enumerate(lines)
            if self._normalize_cell(line.text if isinstance(line, _OCRLine) else line)
        ]

        pending: dict[str, Any] | None = None
        for line in normalized_lines:
            if self._is_summary_line(line.text):
                skipped += 1
                continue

            tx_date = self._parse_date(line.text)
            amount_tokens = self._extract_ocr_amount_tokens(line.text)
            desc_part = self._strip_ocr_date_and_amounts(line.text, amount_tokens)

            if tx_date is not None:
                finalized = self._finalize_ocr_candidate(pending, bank=bank)
                if finalized is not None:
                    items.append(finalized)
                    valid += 1  # type: ignore
                elif pending is not None:
                    skipped += 1  # type: ignore

                pending = {
                    "date": tx_date,
                    "descriptions": [desc_part] if desc_part else [],
                    "amount_tokens": amount_tokens,
                    "confidences": [line.confidence],
                }
                continue

            if pending is None:
                skipped += 1
                continue

            if desc_part:
                pending["descriptions"].append(desc_part)
            if amount_tokens:
                pending["amount_tokens"].extend(amount_tokens)
            pending["confidences"].append(line.confidence)

        finalized = self._finalize_ocr_candidate(pending, bank=bank)
        if finalized is not None:
            items.append(finalized)
            valid += 1  # type: ignore
        elif pending is not None:
            skipped += 1  # type: ignore

        dedup: dict[tuple[str, str, int, str], ParsedTransaction] = {}
        for tx in items:
            key = (tx.date.isoformat(), tx.description.lower(), int(round(tx.amount * 100)), tx.type)
            dedup[key] = tx
        return list(dedup.values()), skipped, valid

    def _extract_ocr_amount_tokens(self, text: str) -> list[str]:
        cleaned = text
        date_token = DATE_CANDIDATE.search(cleaned)
        if date_token:
            cleaned = cleaned.replace(date_token.group(0), " ")

        tokens: list[str] = []
        # Only extract amounts that have explicit currency markers or decimal points
        # to avoid picking up transaction reference numbers
        for match in re.finditer(r"(?:[+-]|\()?\s*(?:₹|INR|RS\.?)?\s*(?:\d{1,3}(?:,\d{2,3})+|\d+)(?:\.\d{1,2})?\)?", cleaned, re.IGNORECASE):
            token = self._normalize_cell(match.group(0))
            if self._looks_like_ocr_amount_token(token):
                tokens.append(token)
        return tokens

    def _looks_like_ocr_amount_token(self, token: str) -> bool:
        if not token:
            return False
        # More lenient: accept anything that looks like a monetary value
        # (has decimal point or thousands separator or currency marker)
        has_decimal_or_separator = any(marker in token for marker in (".", ",", "₹", "INR", "RS"))
        if not has_decimal_or_separator:
            return False
        # Simple check: can we parse it as an amount?
        parsed = self._parse_amount(token)
        return parsed is not None and parsed > 0

    def _strip_ocr_date_and_amounts(self, text: str, amount_tokens: list[str]) -> str:
        desc = text
        date_token = DATE_CANDIDATE.search(desc)
        if date_token:
            desc = desc.replace(date_token.group(0), " ")
        for token in amount_tokens:
            desc = desc.replace(token, " ")
        return re.sub(r"\s+", " ", desc).strip(" -") or ""

    def _finalize_ocr_candidate(self, candidate: dict[str, Any] | None, bank: str = "generic") -> ParsedTransaction | None:
        if not candidate:
            return None

        descriptions = [self._normalize_cell(part) for part in candidate.get("descriptions", []) if self._normalize_cell(part)]
        # Don't require description; if we have date + amount, that's sufficient
        description = " ".join(descriptions).strip()[:255] or "Transaction"

        amount_tokens = candidate.get("amount_tokens", [])
        if not amount_tokens:
            return None

        # Filter to reasonable transaction amounts (< 5 crore = 50 million to exclude balance numbers)
        # Prefer the smallest valid amount in the range
        valid_amounts = []
        for token in amount_tokens:
            amt = self._parse_amount(token)
            # Accept amounts in realistic transaction range (exclude > 50M which are likely balances)
            if amt is not None and amt > 0 and amt < 50000000:  # < 5 crore
                valid_amounts.append((amt, token))

        if not valid_amounts:
            return None

        # Sort by amount value and take the smallest (actual transaction vs balance)
        valid_amounts.sort(key=lambda x: x[0])
        amount_dec = valid_amounts[0][0]
        amount_raw = valid_amounts[0][1]

        balance_dec = None
        if len(valid_amounts) > 1:
            # If we have more than one amount, the largest is likely the balance
            potential_balance = valid_amounts[-1][0]
            if potential_balance > amount_dec:
                balance_dec = potential_balance

        tx_type, tx_amount = self._infer_type_and_amount(
            description,
            None,
            None,
            amount_dec,
            amount_raw,
            bank=bank,
        )
        if tx_type is None or tx_amount is None or tx_amount <= 0:
            return None

        confidence = self._score_ocr_candidate(candidate, balance_dec is not None, tx_type is not None)
        return ParsedTransaction(
            date=candidate["date"],
            description=description,
            amount=tx_amount,
            type=tx_type,
            debit=tx_amount if tx_type == "expense" else None,
            credit=tx_amount if tx_type == "income" else None,
            balance=float(balance_dec) if balance_dec is not None else None,
            confidence=confidence,
            is_valid=True,
        )

    def _score_ocr_candidate(self, candidate: dict[str, Any], has_balance: bool, has_type: bool) -> float:
        confidences = [float(conf) for conf in candidate.get("confidences", []) if conf is not None]
        base = sum(confidences) / len(confidences) if confidences else 0.45
        if not has_balance:
            base -= 0.10
        if len(candidate.get("descriptions", [])) > 2:
            base -= 0.03
        if not has_type:
            base -= 0.10
        return round(max(0.0, min(base, 0.99)), 4)

    def _ocr_type(self, line: str, amount: Decimal | None, bank: str = "generic") -> Literal["income", "expense"] | None:
        inferred = self._infer_type_from_text(line)
        if inferred is not None:
            return inferred
        if bank == "sbi" and amount is not None:
            return "expense" if amount < 0 else "income"
        if amount is not None and amount < 0:
            return "expense"
        return None

    def _clean_with_polars(self, rows: list[ParsedTransaction]) -> list[ParsedTransaction]:
        if not rows or pl is None:
            return rows
        df = pl.DataFrame(
            {
                "description": [r.description for r in rows],
            }
        )
        cleaned = (
            df.with_columns(
                pl.col("description")
                .str.replace_all(r"(?i)\bupi[-\w@./]+\b", " ")
                .str.replace_all(r"(?i)\bpos[ -]?\d+\b", " ")
                .str.replace_all(r"\b\d{6,}\b", " ")
                .str.replace_all(r"(?i)\b(?:txn|ref|utr|id|no|number)\b", " ")
                .str.replace_all(r"\s+", " ")
                .str.strip_chars()
                .alias("description")
            )
            .to_dict(as_series=False)
            .get("description", [])
        )
        normalized: list[ParsedTransaction] = []
        for idx, tx in enumerate(rows):
            desc = cleaned[idx].strip() if idx < len(cleaned) else tx.description
            normalized.append(tx.model_copy(update={"description": (desc or "Transaction")[:255]}))
        return normalized

    def verify_integrity(self, transactions: list[ParsedTransaction]) -> tuple[list[ParsedTransaction], float]:
        if not transactions:
            return transactions, 1.0

        ordered = sorted(transactions, key=lambda t: (t.date.isoformat(), t.description.lower(), t.amount, t.type))
        initial = self._infer_initial_balance(ordered)
        if initial is None:
            return ordered, 1.0

        running = initial
        validated: list[ParsedTransaction] = []
        valid_rows = 0
        for tx in ordered:
            if tx.type == "income":
                running += float(tx.amount)
            else:
                running -= float(tx.amount)

            is_valid = True
            if tx.balance is not None:
                is_valid = abs(running - float(tx.balance)) <= 1.0
            if is_valid:
                valid_rows += 1  # type: ignore
            validated.append(tx.model_copy(update={"is_valid": is_valid}))

        score = valid_rows / len(validated) if validated else 1.0  # type: ignore
        return validated, score

    def _infer_initial_balance(self, ordered: list[ParsedTransaction]) -> float | None:
        for tx in ordered:
            if tx.balance is None:
                continue
            if tx.type == "income":
                return float(tx.balance) - float(tx.amount)
            return float(tx.balance) + float(tx.amount)
        return None

    def _cell(self, row: list[str | None], idx: int | None) -> str:
        if idx is None or idx >= len(row):
            return ""
        return self._normalize_cell(row[idx])

    def _normalize_cell(self, value: str | None) -> str:
        return re.sub(r"\s+", " ", str(value or "").replace("\u00a0", " ")).strip()

    def _decode_text(self, file_bytes: bytes) -> str:
        for encoding in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                return file_bytes.decode(encoding)
            except Exception:
                continue
        return ""

    def _is_csv_input(self, file_bytes: bytes, filename: str | None, content_type: str | None) -> bool:
        name = (filename or "").lower().strip()
        ctype = (content_type or "").lower().strip()
        if name.endswith(".csv"):
            return True
        if "csv" in ctype:
            return True

        text = self._decode_text(file_bytes[:4096])
        if not text:
            return False
        first_line = text.splitlines()[0].lower() if text.splitlines() else ""
        csv_signals = ("date", "description", "narration", "debit", "credit", "amount", "balance")
        return "," in first_line and any(sig in first_line for sig in csv_signals)

    def _is_xlsx_input(self, file_bytes: bytes, filename: str | None, content_type: str | None) -> bool:
        name = (filename or "").lower().strip()
        ctype = (content_type or "").lower().strip()
        if name.endswith(".xlsx"):
            return True
        if "spreadsheetml" in ctype or "excel" in ctype:
            return True
        return False

    def _tx_key(self, tx: ParsedTransaction) -> tuple[str, str, int, str]:
        return (tx.date.isoformat(), tx.description.lower(), int(round(float(tx.amount) * 100)), tx.type)

    def _is_summary_line(self, text: str) -> bool:
        value = self._normalize_cell(text).lower()
        return any(pattern.search(value) for pattern in SUMMARY_PATTERNS)

    def _lookup_csv_field(self, row: dict[str, str], aliases: tuple[str, ...]) -> str:
        normalized = {self._normalize_cell(k).lower(): self._normalize_cell(v) for k, v in row.items()}
        for alias in aliases:
            if alias in normalized:
                return normalized[alias]

        for key, value in normalized.items():
            key_tokens = {token for token in re.split(r"[^a-z0-9]+", key) if token}
            for alias in aliases:
                alias_tokens = [token for token in re.split(r"[^a-z0-9]+", alias.lower()) if token]
                if alias_tokens and all(token in key_tokens for token in alias_tokens):
                    return value
        return ""

    def _has_explicit_sign(self, raw_amount: str | None) -> bool:
        value = self._normalize_cell(raw_amount)
        if not value:
            return False
        return value.startswith(("+", "-")) or (value.startswith("(") and value.endswith(")"))

    def _infer_type_from_text(self, text: str) -> Literal["income", "expense"] | None:
        upper = (text or "").upper()
        expense_hits = any(k in upper for k in ("OUTWARD", "UPI", "POS", "DEBIT", "DR", "WITHDRAW", "PURCHASE"))
        income_hits = any(k in upper for k in ("INWARD", "CR", "SALARY", "CREDIT", "DEPOSIT", "REFUND"))
        if income_hits and not expense_hits:
            return "income"
        if expense_hits and not income_hits:
            return "expense"
        return None

    def _infer_type_and_amount(
        self,
        raw_desc: str,
        debit: Decimal | None,
        credit: Decimal | None,
        amount: Decimal | None,
        raw_amount: str | None = None,
        bank: str = "generic",
    ) -> tuple[Literal["income", "expense"] | None, float | None]:
        if bank == "axis":
            if debit is not None and debit > 0:
                return "expense", float(abs(debit))
            if credit is not None and credit > 0:
                return "income", float(abs(credit))

        if debit is not None and debit > 0:
            return "expense", float(abs(debit))
        if credit is not None and credit > 0:
            return "income", float(abs(credit))
        if amount is None or amount == 0:
            return None, None

        # Reject unreasonably large amounts (likely balance values, not transactions)
        # Use a more lenient threshold of 50 million (5 crore)
        abs_amount = float(abs(amount))
        if abs_amount >= 50000000:  # 5 crore - balance threshold
            return None, None

        if bank == "sbi" and self._has_explicit_sign(raw_amount):
            return ("expense" if amount < 0 else "income"), float(abs(amount))

        if self._has_explicit_sign(raw_amount):
            return ("expense" if amount < 0 else "income"), float(abs(amount))

        inferred = self._infer_type_from_text(raw_desc)
        if inferred is None and amount < 0:
            inferred = "expense"
        if inferred is None:
            return None, None
        return inferred, float(abs(amount))

    def _parse_csv_row(self, row: dict[str, str], bank: str = "generic") -> ParsedTransaction | None:
        raw_date = self._lookup_csv_field(row, DATE_ALIASES)
        raw_desc = self._row_description_from_csv(row, bank=bank)
        if not raw_date or not raw_desc or self._is_summary_line(raw_desc):
            return None

        tx_date = self._parse_date(raw_date)
        if tx_date is None:
            return None

        raw_amount = self._lookup_csv_field(row, AMOUNT_ALIASES)
        debit = self._parse_amount(self._lookup_csv_field(row, DEBIT_ALIASES))
        credit = self._parse_amount(self._lookup_csv_field(row, CREDIT_ALIASES))
        amount = self._parse_amount(raw_amount)
        balance = self._parse_amount(self._lookup_csv_field(row, BALANCE_ALIASES))

        tx_type, tx_amount = self._infer_type_and_amount(raw_desc, debit, credit, amount, raw_amount, bank=bank)
        if tx_type is None or tx_amount is None or tx_amount <= 0:
            return None

        return ParsedTransaction(
            date=tx_date,
            description=raw_desc.strip()[:255] or "Transaction",
            amount=tx_amount,
            type=tx_type,
            balance=float(balance) if balance is not None else None,
            is_valid=True,
        )

    def _row_description_from_csv(self, row: dict[str, str], bank: str = "generic") -> str:
        desc = self._lookup_csv_field(row, DESCRIPTION_ALIASES)
        if bank != "hdfc":
            return desc

        normalized = {self._normalize_cell(k).lower(): self._normalize_cell(v) for k, v in row.items()}
        parts: list[str] = []
        seen: set[str] = set()
        for key, value in normalized.items():
            if not value:
                continue
            key_tokens = {token for token in re.split(r"[^a-z0-9]+", key) if token}
            if key_tokens & {"date", "debit", "withdrawal", "credit", "deposit", "amount", "balance"}:
                continue
            if self._is_monetary_cell(value) or value in seen:
                continue
            seen.add(value)
            parts.append(value)
        return " ".join(parts).strip() or desc

    def _parse_date(self, value: str) -> date | None:
        candidate_match = DATE_CANDIDATE.search(value)
        candidate = candidate_match.group(0) if candidate_match else value.strip()
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y", "%Y-%m-%d", "%d %b %Y", "%d %b %y", "%d %B %Y"):
            try:
                return datetime.strptime(candidate, fmt).date()
            except ValueError:
                continue
        return None

    def _parse_amount(self, raw: str) -> Decimal | None:
        value = (raw or "").strip()
        if not value:
            return None
        negative = value.startswith("(") and value.endswith(")")
        value = value.replace("₹", "").replace("INR", "").replace("Rs.", "").replace("Rs", "").strip()
        value = value.strip("()")
        if not value:
            return None
        match = AMOUNT_CANDIDATE.search(value)
        if not match:
            return None
        token = match.group(0).replace(",", "")
        try:
            amount = Decimal(token)
            return -amount if negative else amount
        except InvalidOperation:
            return None

    def _is_monetary_cell(self, raw: str | None) -> bool:
        value = self._normalize_cell(raw)
        if not value or re.search(r"[A-Za-z]", value):
            return False
        if not MONEY_LIKE_CELL.match(value):
            return False
        digits_only = re.sub(r"\D", "", value)
        has_currency_markers = any(char in value for char in (".", ",", "(", ")", "+", "-", "₹"))
        if not has_currency_markers and len(digits_only) >= 8:
            return False
        return self._parse_amount(value) is not None


class LLMParser:
    """
    Production-grade LLM parser backed by a local Ollama model.
    Handles unstructured documents by chunking text and extracting transactions as JSON.
    """

    _CHUNK_SIZE = 1200
    _MAX_CHUNKS = 8
    _JSON_RETRIES = 3
    _TIMEOUT = 120.0

    _PROMPT_TEMPLATE = (
        "You are a financial data extraction assistant.\n"
        "Extract ALL bank transactions from the text below.\n"
        "Return ONLY a valid JSON array (no markdown, no explanation).\n"
        "Each element must have these exact keys:\n"
        '  "date" (YYYY-MM-DD), "description" (string), "amount" (number > 0),\n'
        '  "type" ("income" or "expense"), "balance" (number or null),\n'
        '  "debit" (number or null), "credit" (number or null)\n'
        "Do NOT invent values. Use null for unknown fields.\n\n"
        "TEXT:\n{chunk}\n\n"
        "JSON ARRAY:"
    )

    def __init__(self) -> None:
        from app.core.config import settings
        self.settings = settings
        self._base_url = settings.ollama_base_url
        self._primary = settings.ollama_primary_model
        self._fallback = settings.ollama_fallback_model

    def parse(self, text: str, filename: str | None = None) -> ParserResponse:
        logger.info("llm_parse_start file=%s", filename)

        sanitized = self._sanitize_text(text)
        if not sanitized:
            return ParserResponse(
                status="empty_content",
                reconciliation_score=1.0,
                transactions=[],
                meta={"method": "llm_local", "notice": "No extractable text found"}
            )

        chunks = self._chunk_text(sanitized)
        logger.info("llm_parse chunks=%d file=%s", len(chunks), filename)

        # Try primary model
        transactions, model_used = self._parse_with_model(chunks, self._primary)

        # Fallback if primary produced nothing
        if not transactions and self._fallback and self._fallback != self._primary:
            logger.info("llm_parse primary_empty_fallback model=%s", self._fallback)
            transactions, model_used = self._parse_with_model(chunks, self._fallback)

        confidence = self._compute_confidence(transactions, len(chunks))

        logger.info(
            "llm_parse_done file=%s model=%s txns=%d confidence=%.4f",
            filename, model_used, len(transactions), confidence
        )

        return ParserResponse(
            status="success" if transactions else "no_transactions",
            reconciliation_score=round(confidence, 4),
            transactions=transactions,
            meta={
                "method": "llm_local",
                "model_used": model_used,
                "confidence_score": round(confidence, 4),
                "count": len(transactions),
                "chunks_processed": len(chunks)
            }
        )

    def _parse_with_model(self, chunks: list[str], model: str) -> tuple[list[ParsedTransaction], str]:
        all_txns: list[ParsedTransaction] = []
        seen_keys: set[tuple] = set()

        for chunk in chunks:
            prompt = self._PROMPT_TEMPLATE.format(chunk=chunk)
            raw_json = self._call_ollama_with_retry(prompt, model)
            if not raw_json:
                continue

            for tx in self._parse_raw_json(raw_json):
                # Simple dedup within the session
                key = (tx.date.isoformat(), tx.description.lower()[:60], round(float(tx.amount), 2))
                if key not in seen_keys:
                    seen_keys.add(key)
                    all_txns.append(tx)

        return all_txns, model

    def _call_ollama_with_retry(self, prompt: str, model: str) -> str | None:
        for attempt in range(1, self._JSON_RETRIES + 1):
            raw = self._call_ollama(prompt, model)
            if raw is None:
                continue

            cleaned = self._clean_json(raw)
            if self._is_valid_json_array(cleaned):
                return cleaned

            logger.debug("ollama_retry attempt=%d model=%s invalid_json", attempt, model)

        logger.debug("ollama_retry_failed model=%s", model)
        return None

    def _call_ollama(self, prompt: str, model: str) -> str | None:
        url = f"{self._base_url}/api/generate"
        payload = {
            "model": model,
            "prompt": prompt,
            "stream": False,
            "options": {"temperature": 0.1, "num_predict": 2048}
        }
        try:
            logger.debug("ollama_request model=%s prompt=%s", model, prompt[:100])
            with httpx.Client(timeout=self._TIMEOUT) as client:
                resp = client.post(url, json=payload)
                resp.raise_for_status()
                response_text = resp.json().get("response")
                logger.debug("ollama_response_raw model=%s text=%s", model, response_text[:200] if response_text else "None")
                return response_text
        except Exception as e:
            logger.error("ollama_error model=%s error=%s", model, str(e))
            return None

    def _parse_raw_json(self, raw_json: str) -> list[ParsedTransaction]:
        try:
            logger.debug("ollama_parsing_json raw=%s", raw_json[:200])
            items = json.loads(raw_json)
            if not isinstance(items, list):
                if isinstance(items, dict) and "transactions" in items:
                    items = items["transactions"]
                else:
                    logger.debug("ollama_parse_json_error raw=%s error=%s", raw_json[:200], "Not a list or dict with 'transactions'")
                    return []
        except (json.JSONDecodeError, ValueError) as e:
            logger.debug("ollama_parse_json_error raw=%s error=%s", raw_json[:200], e)
            return []

        parsed: list[ParsedTransaction] = []
        for item in items:
            if not isinstance(item, dict):
                logger.debug("ollama_parse_json_item_error item=%s error=%s", str(item)[:100], "Not a dictionary")
                continue
            try:
                # Basic normalization
                if "date" in item and isinstance(item["date"], str):
                    # Pydantic handles the rest
                    parsed.append(ParsedTransaction(**{k: v for k, v in item.items() if k in ParsedTransaction.model_fields}))
            except Exception as e:
                logger.debug("ollama_parse_json_item_error item=%s error=%s", str(item)[:100], e)
                continue
        return parsed

    def _chunk_text(self, text: str) -> list[str]:
        if len(text) <= self._CHUNK_SIZE:
            return [text]

        chunks: list[str] = []
        start = 0
        while start < len(text) and len(chunks) < self._MAX_CHUNKS:
            end = min(start + self._CHUNK_SIZE, len(text))
            if end < len(text):
                # Avoid splitting mid-line
                newline = text.rfind("\n", start, end)
                if newline > start:
                    end = newline + 1
            chunks.append(text[start:end])
            start = end
        return chunks

    def _compute_confidence(self, transactions: list[ParsedTransaction], n_chunks: int) -> float:
        """
        Confidence (0.0 - 1.0) based on:
        - JSON validity (implied if we got txns) : 0.4
        - Extraction density (txns / chunk)      : 0.3
        - Field ratio (missing fields)           : 0.3
        """
        if not transactions:
            return 0.0

        # Field completeness
        total_fields = len(transactions) * 4  # date, desc, amt, type
        filled_fields = sum(
            (1 if tx.date else 0) + (1 if tx.description else 0) + (1 if tx.amount > 0 else 0) + (1 if tx.type else 0)
            for tx in transactions
        )
        field_ratio = filled_fields / total_fields if total_fields > 0 else 0

        # Density: at least 1 txn per chunk is good
        density_ratio = min(1.0, len(transactions) / max(1, n_chunks))

        score = 0.4 + (density_ratio * 0.3) + (field_ratio * 0.3)
        return min(1.0, score)

    def _sanitize_text(self, text: str) -> str:
        if not text:
            return ""
        cleaned = text.replace("\x00", " ")
        cleaned = re.sub(r"\s+", " ", cleaned).strip()
        return cleaned[:15000]

    def _clean_json(self, raw: str) -> str:
        if not raw:
            return "[]"
        # Strip <think> tags from deepseek models
        raw = re.sub(r"<think>[\s\S]*?</think>", "", raw, flags=re.IGNORECASE).strip()
        match = re.search(r"```(?:json)?\s*([\s\S]+?)\s*```", raw)
        if match:
            return match.group(1).strip()
        start = raw.find("[")
        end = raw.rfind("]")
        if start != -1 and end > start:
            return raw[start: end + 1].strip()
        return raw.strip()

    def _is_valid_json_array(self, s: str) -> bool:
        try:
            val = json.loads(s)
            return isinstance(val, (list, dict))
        except (json.JSONDecodeError, ValueError):
            return False
