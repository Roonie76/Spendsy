from __future__ import annotations

import csv
import io
import logging
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Literal

import pdfplumber
try:
    import openpyxl
except ImportError:
    openpyxl = None
try:
    import polars as pl
except ImportError:
    pl = None

from app.core.base_parser import BaseParser
from app.core.schemas import ParsedTransaction, ParserResponse

logger = logging.getLogger(__name__)

# Constants and Regexes from legacy parser.py
DATE_CANDIDATE = re.compile(r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b")
AMOUNT_CANDIDATE = re.compile(r"-?(?:\d{1,3}(?:,\d{2,3})+|\d+)(?:\.\d{1,2})?")
MONEY_LIKE_CELL = re.compile(
    r"^\s*(?:[+-]|\()?\s*(?:₹|INR|RS\.?)?\s*(?:\d{1,3}(?:,\d{2,3})+|\d+)(?:\.\d{1,2})?\s*\)?\s*$",
    re.IGNORECASE,
)
SUMMARY_PATTERNS = (
    re.compile(r"^\s*(opening|closing)\s+balance\b", re.IGNORECASE),
    re.compile(r"^\s*(b/f|brought forward)\b", re.IGNORECASE),
    re.compile(r"^\s*total\s+(debit|credit|withdraw|deposit|amount|txn|transactions)\b", re.IGNORECASE),
)
DATE_ALIASES = ("date", "txn date", "transaction date", "value date", "tran date")
DESCRIPTION_ALIASES = (
    "description", "narration", "particulars", "details", "remarks",
    "transaction remarks", "transaction details", "ref particulars",
)
DEBIT_ALIASES = ("debit", "withdrawal", "dr", "amt out", "payment", "withdrawal amt", "debit amount")
CREDIT_ALIASES = ("credit", "deposit", "cr", "amt in", "receipt", "deposit amt", "credit amount")
AMOUNT_ALIASES = ("amount", "txn amount", "transaction amount")
BALANCE_ALIASES = ("balance", "running balance", "closing balance", "available balance")

@dataclass(frozen=True)
class _HeaderMap:
    date_idx: int
    desc_idx: int
    debit_idx: int | None
    credit_idx: int | None
    amount_idx: int | None
    balance_idx: int | None

class TabularParser(BaseParser):
    """
    Handles structured tabular data from CSV, XLSX, and digital PDF tables.
    Refactored from legacy IntegratedParser to focus only on deterministic tabular extraction.
    """
    @property
    def name(self) -> str:
        return "tabular"

    @property
    def version(self) -> str:
        return "2.1.0"

    @property
    def priority(self) -> int:
        return 10

    def can_handle(self, content: bytes, text: str, **kwargs: Any) -> float:
        filename = str(kwargs.get("filename", "")).lower()
        content_type = str(kwargs.get("content_type", "")).lower()
        
        if filename.endswith((".csv", ".xlsx", ".xls")) or "spreadsheet" in content_type or "csv" in content_type:
            return 1.0
            
        if filename.endswith(".pdf"):
            # Check if text contains table-like structures (e.g., many dates and amounts)
            date_count = len(DATE_CANDIDATE.findall(text))
            if date_count > 10:
                return 0.85
            return 0.6
            
        return 0.1

    def __init__(self) -> None:
        pass

    def parse(self, content: bytes, text: str, **kwargs: Any) -> ParserResponse:
        filename = kwargs.get("filename")
        content_type = kwargs.get("content_type")
        parsed_rows: list[ParsedTransaction] = []
        method = "digital"
        bank = kwargs.get("bank", "generic")

        if self._is_csv_input(content, filename, content_type):
            parsed_rows = self._extract_csv(content, bank=bank)
            method = "csv"
        elif self._is_xlsx_input(content, filename, content_type):
            parsed_rows = self._extract_xlsx(content, bank=bank)
            method = "xlsx"
        else:
            parsed_rows = self._extract_digital(content, bank=bank)
            method = "digital"

        # Basic cleaning (moved here from legacy)
        cleaned = self._clean_descriptions(parsed_rows)
        
        # Sort by date
        ordered = sorted(
            cleaned,
            key=lambda t: (t.date.isoformat(), t.description.lower(), round(float(t.amount), 2), t.type),
        )

        return ParserResponse(
            status="success" if ordered else "no_transactions",
            reconciliation_score=1.0, # Will be computed by the pipeline's ReconciliationEngine
            transactions=ordered,
            meta={
                "method": method,
                "count": len(ordered),
                "parser_name": self.name,
                "parser_version": self.version,
            },
        )

    def _extract_digital(self, pdf_bytes: bytes, bank: str = "generic") -> list[ParsedTransaction]:
        transactions: list[ParsedTransaction] = []
        try:
            with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                for page_idx, page in enumerate(pdf.pages):
                    tables = page.extract_tables() or []
                    for table in tables:
                        transactions.extend(self._parse_structured_table(table, bank=bank))
        except Exception as e:
            logger.error("tabular_parser.extract_digital error=%s", str(e))
        return transactions

    def _extract_xlsx(self, file_bytes: bytes, bank: str = "generic") -> list[ParsedTransaction]:
        if openpyxl is None:
            return []
        transactions: list[ParsedTransaction] = []
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            for sheet in workbook.worksheets:
                rows = [
                    [self._normalize_cell(cell) for cell in row]
                    for row in sheet.iter_rows(values_only=True)
                    if any(self._normalize_cell(cell) for cell in row)
                ]
                if rows:
                    transactions.extend(self._parse_structured_table(rows, bank=bank))
        except Exception as e:
            logger.error("tabular_parser.extract_xlsx error=%s", str(e))
        return transactions

    def _extract_csv(self, file_bytes: bytes, bank: str = "generic") -> list[ParsedTransaction]:
        rows: list[dict[str, str]] = []
        if pl is not None:
            try:
                df = pl.read_csv(io.BytesIO(file_bytes), ignore_errors=True, infer_schema_length=5000)
                rows = [{k: str(v or "") for k, v in row.items()} for row in df.iter_rows(named=True)]
            except Exception:
                rows = []
        if not rows:
            text = self._decode_text(file_bytes)
            if text:
                reader = csv.DictReader(io.StringIO(text))
                rows = [{str(k or ""): str(v or "") for k, v in row.items()} for row in reader]

        parsed: list[ParsedTransaction] = []
        for row in rows:
            tx = self._parse_csv_row(row, bank=bank)
            if tx:
                parsed.append(tx)
        return parsed

    def _parse_structured_table(self, table: list[list[str | None]], bank: str = "generic") -> list[ParsedTransaction]:
        if not table: return []
        header = self._detect_header(table)
        if header is None: return []

        transactions: list[ParsedTransaction] = []
        # Coalesce multi-line rows
        coalesced = self._coalesce_table_rows(table[1:], header, bank=bank)
        for row in coalesced:
            tx = self._parse_table_row(row, header, bank=bank)
            if tx:
                transactions.append(tx)
        return transactions

    def _detect_header(self, table: list[list[str | None]]) -> _HeaderMap | None:
        for row in table[:4]:
            norm = [self._normalize_cell(c) for c in row]
            date_idx = self._find_col(norm, DATE_ALIASES)
            desc_idx = self._find_col(norm, DESCRIPTION_ALIASES)
            debit_idx = self._find_col(norm, DEBIT_ALIASES)
            credit_idx = self._find_col(norm, CREDIT_ALIASES)
            amount_idx = self._find_col(norm, AMOUNT_ALIASES)
            balance_idx = self._find_col(norm, BALANCE_ALIASES)
            if date_idx is not None and desc_idx is not None and (amount_idx is not None or debit_idx is not None or credit_idx is not None):
                return _HeaderMap(date_idx, desc_idx, debit_idx, credit_idx, amount_idx, balance_idx)
        return None

    def _find_col(self, headers: list[str], aliases: tuple[str, ...]) -> int | None:
        for idx, h in enumerate(headers):
            if any(a.lower() in h.lower() for a in aliases):
                return idx
        return None

    def _coalesce_table_rows(self, rows: list[list[str | None]], hm: _HeaderMap, bank: str = "generic") -> list[list[str | None]]:
        merged: list[list[str | None]] = []
        for raw_row in rows:
            row = list(raw_row or [])
            if merged and self._is_continuation(row, merged[-1], hm, bank):
                prev = merged[-1]
                extra_desc = self._continuation_text(row, hm, bank)
                if extra_desc:
                    if hm.desc_idx < len(prev):
                        prev[hm.desc_idx] = f"{prev[hm.desc_idx] or ''} {extra_desc}".strip()
                # Carry forward amounts/balances if they appeared on the second line
                self._carry_forward(prev, row, hm)
                continue
            merged.append(row)
        return merged

    def _is_continuation(self, row: list[str | None], prev: list[str | None], hm: _HeaderMap, bank: str) -> bool:
        if row and self._normalize_cell(row[0]): return False # First cell usually has date
        if self._parse_date(self._cell(row, hm.date_idx)): return False
        desc = self._continuation_text(row, hm, bank)
        if desc and any(p.search(desc.lower()) for p in SUMMARY_PATTERNS): return False
        if any(idx is not None and self._is_monetary_cell(self._cell(row, idx)) for idx in (hm.debit_idx, hm.credit_idx, hm.amount_idx)):
            return False
        return bool(desc)

    def _continuation_text(self, row: list[str | None], hm: _HeaderMap, bank: str) -> str:
        parts = []
        reserved = {hm.date_idx}
        for idx, val in enumerate(row):
            cell = self._normalize_cell(val)
            if not cell or idx in reserved or self._is_monetary_cell(cell): continue
            parts.append(cell)
        return " ".join(parts).strip()

    def _carry_forward(self, prev: list[str | None], row: list[str | None], hm: _HeaderMap) -> None:
        for idx in (hm.debit_idx, hm.credit_idx, hm.amount_idx, hm.balance_idx):
            if idx is not None and idx < len(row):
                val = self._cell(row, idx)
                if val and self._is_monetary_cell(val) and not self._cell(prev, idx):
                    if idx >= len(prev): prev.extend([""] * (idx - len(prev) + 1))
                    prev[idx] = val

    def _parse_table_row(self, row: list[str | None], hm: _HeaderMap, bank: str) -> ParsedTransaction | None:
        raw_date = self._cell(row, hm.date_idx)
        raw_desc = self._continuation_text(row, hm, bank) if bank == "hdfc" else self._cell(row, hm.desc_idx)
        if not raw_date or not raw_desc or any(p.search(raw_desc.lower()) for p in SUMMARY_PATTERNS):
            return None
        
        tx_date = self._parse_date(raw_date)
        if not tx_date: return None

        debit = self._parse_amount(self._cell(row, hm.debit_idx))
        credit = self._parse_amount(self._cell(row, hm.credit_idx))
        amount = self._parse_amount(self._cell(row, hm.amount_idx))
        balance = self._parse_amount(self._cell(row, hm.balance_idx))

        tx_type, tx_amount = self._infer_type_amount(raw_desc, debit, credit, amount, self._cell(row, hm.amount_idx), bank)
        if not tx_type or not tx_amount: return None

        return ParsedTransaction(
            date=tx_date,
            description=raw_desc[:255],
            amount=tx_amount,
            type=tx_type,
            debit=tx_amount if tx_type == "expense" else None,
            credit=tx_amount if tx_type == "income" else None,
            balance=float(balance) if balance is not None else None,
        )

    def _parse_csv_row(self, row: dict[str, str], bank: str) -> ParsedTransaction | None:
        raw_date = self._lookup_csv(row, DATE_ALIASES)
        tx_date = self._parse_date(raw_date)
        if not tx_date: return None

        raw_desc = self._lookup_csv(row, DESCRIPTION_ALIASES)
        if not raw_desc or any(p.search(raw_desc.lower()) for p in SUMMARY_PATTERNS): return None

        debit = self._parse_amount(self._lookup_csv(row, DEBIT_ALIASES))
        credit = self._parse_amount(self._lookup_csv(row, CREDIT_ALIASES))
        amount = self._parse_amount(self._lookup_csv(row, AMOUNT_ALIASES))
        balance = self._parse_amount(self._lookup_csv(row, BALANCE_ALIASES))

        tx_type, tx_amount = self._infer_type_amount(raw_desc, debit, credit, amount, self._lookup_csv(row, AMOUNT_ALIASES), bank)
        if not tx_type or not tx_amount: return None

        return ParsedTransaction(
            date=tx_date,
            description=raw_desc[:255],
            amount=tx_amount,
            type=tx_type,
            balance=float(balance) if balance is not None else None,
        )

    def _lookup_csv(self, row: dict[str, str], aliases: tuple[str, ...]) -> str:
        norm_row = {self._normalize_cell(k).lower(): v for k, v in row.items()}
        for a in aliases:
            if a.lower() in norm_row: return self._normalize_cell(norm_row[a.lower()])
        return ""

    def _infer_type_amount(self, desc: str, dr: Decimal | None, cr: Decimal | None, amt: Decimal | None, raw_amt: str, bank: str) -> tuple[str | None, float | None]:
        if dr and dr > 0: return "expense", float(dr)
        if cr and cr > 0: return "income", float(cr)
        if not amt: return None, None
        
        abs_amt = float(abs(amt))
        if abs_amt > 50000000: return None, None # Likely a balance

        if raw_amt.startswith(("-", "(")) or "dr" in desc.lower() or "debit" in desc.lower():
            return "expense", abs_amt
        if raw_amt.startswith("+") or "cr" in desc.lower() or "credit" in desc.lower():
            return "income", abs_amt
        
        # Heuristic fallback
        if amt < 0: return "expense", abs_amt
        # Use simple keyword check for income
        if any(k in desc.upper() for k in ("SALARY", "REFUND", "INTEREST")): return "income", abs_amt
        return "expense", abs_amt

    def _clean_descriptions(self, rows: list[ParsedTransaction]) -> list[ParsedTransaction]:
        # Simple non-polars cleaning
        for tx in rows:
            d = tx.description
            d = re.sub(r"(?i)\bupi[-\w@./]+\b", " ", d)
            d = re.sub(r"\b\d{6,}\b", " ", d)
            tx.description = re.sub(r"\s+", " ", d).strip() or "Transaction"
        return rows

    def _parse_date(self, value: str) -> date | None:
        match = DATE_CANDIDATE.search(value)
        cand = match.group(0) if match else value.strip()
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y", "%Y-%m-%d", "%d %b %Y"):
            try: return datetime.strptime(cand, fmt).date()
            except ValueError: continue
        return None

    def _parse_amount(self, raw: str) -> Decimal | None:
        v = (raw or "").strip().replace(",", "").replace("₹", "").replace("INR", "").replace("Rs", "").strip("()")
        match = AMOUNT_CANDIDATE.search(v)
        if not match: return None
        try: return Decimal(match.group(0))
        except InvalidOperation: return None

    def _is_monetary_cell(self, raw: str | None) -> bool:
        v = self._normalize_cell(raw)
        if not v or re.search(r"[A-Za-z]", v): return False
        return self._parse_amount(v) is not None

    def _normalize_cell(self, v: Any) -> str:
        return re.sub(r"\s+", " ", str(v or "").replace("\u00a0", " ")).strip()

    def _decode_text(self, b: bytes) -> str:
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try: return b.decode(enc)
            except Exception: continue
        return ""

    def _cell(self, row: list[str | None], idx: int | None) -> str:
        if idx is None or idx >= len(row): return ""
        return self._normalize_cell(row[idx])

    def _is_csv_input(self, b: bytes, fn: str | None, ct: str | None) -> bool:
        if fn and fn.lower().endswith(".csv"): return True
        if ct and "csv" in ct.lower(): return True
        return "," in self._decode_text(b[:1024])

    def _is_xlsx_input(self, b: bytes, fn: str | None, ct: str | None) -> bool:
        if fn and fn.lower().endswith(".xlsx"): return True
        return ct is not None and ("excel" in ct.lower() or "spreadsheet" in ct.lower())
